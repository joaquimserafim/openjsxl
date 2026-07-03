#!/usr/bin/env python3
"""Out-of-band reference benchmarks for the Python ecosystem (openpyxl, python-calamine).

These are the numbers openjsxl measures itself against — Python's calamine binding is the speed
bar, openpyxl the capability bar (see the repo README). They live outside the JS harness because
they need a Python environment, not npm; the JS `pnpm bench` run is self-contained without them.

Run AFTER `pnpm bench` (or `pnpm bench --quick`) has authored the read fixtures under
../.cache/. This script reads those exact files, so the read comparison is truly like-for-like.

    python3 -m venv .venv && . .venv/bin/activate
    pip install -r packages/bench/py/requirements.txt
    python3 packages/bench/py/bench_py.py            # -> packages/bench/.cache/python.json
    pnpm bench                                        # re-render docs with the Python column filled

Each (library, op, workload, size) cell runs in its own subprocess (mirroring the JS harness), so
`ru_maxrss` is that operation's peak, uncontaminated by the others.
"""

import argparse
import json
import os
import resource
import statistics
import subprocess
import sys
import time
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "..", ".cache")
COLS = 10
SIZES = [("10k", 1_000), ("100k", 10_000), ("1M", 100_000)]
WORDS = [
    "Alpha", "Bravo", "Charlie", "Delta", "Echo", "Foxtrot", "Golf", "Hotel",
    "India", "Juliet", "Kilo", "Lima", "Mike", "November", "Oscar", "Papa",
    "Quebec", "Romeo", "Sierra", "Tango", "Uniform", "Victor", "Whiskey", "Xray",
    "Yankee", "Zulu",
]


def peak_rss_mb():
    """Max resident set of this process, in MiB (ru_maxrss is bytes on macOS, KiB on Linux)."""
    ru = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    return (ru if sys.platform == "darwin" else ru * 1024) / 1024 / 1024


def cell_value(kind, r, c):
    if kind == "strings":
        return f"{WORDS[(r * COLS + c) % len(WORDS)]}-{r % 500}"
    return round((r * COLS + c) * 150 + c * 7 + 0.25) / 100


def dataset(kind, rows):
    return [[cell_value(kind, r, c) for c in range(COLS)] for r in range(rows)]


def fixture_path(workload, size_key):
    return os.path.join(CACHE, f"read-{workload}-{size_key}.xlsx")


# --- operations -----------------------------------------------------------------------------

def read_openpyxl(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sink = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, (int, float)):
                    sink += v
                elif isinstance(v, str):
                    sink += len(v)
    wb.close()
    return sink


def read_calamine(path):
    from python_calamine import CalamineWorkbook

    wb = CalamineWorkbook.from_path(path)
    sink = 0
    for name in wb.sheet_names:
        for row in wb.get_sheet_by_name(name).to_python():
            for v in row:
                if isinstance(v, (int, float)):
                    sink += v
                elif isinstance(v, str):
                    sink += len(v)
    return sink


def write_openpyxl(kind, rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Bench")
    styled = kind == "styled"
    fills = [PatternFill("solid", fgColor="DDEBF7"), None, PatternFill("solid", fgColor="FCE4D6")]
    fonts = [Font(bold=True, color="1F4E79"), Font(italic=True), Font(bold=True)]
    from openpyxl.cell import WriteOnlyCell

    for r in range(rows):
        if not styled:
            ws.append([cell_value(kind, r, c) for c in range(COLS)])
        else:
            cells = []
            for c in range(COLS):
                cell = WriteOnlyCell(ws, value=cell_value("numbers", r, c))
                cell.font = fonts[c % 3]
                if fills[c % 3] is not None:
                    cell.fill = fills[c % 3]
                cells.append(cell)
            ws.append(cells)
    out = os.path.join(CACHE, f"_py_write_{kind}.xlsx")
    wb.save(out)
    size = os.path.getsize(out)
    os.remove(out)
    return size


OPS = {
    ("read", "openpyxl"): lambda cfg: read_openpyxl(fixture_path(cfg["workload"], cfg["sizeKey"])),
    ("read", "python-calamine"): lambda cfg: read_calamine(fixture_path(cfg["workload"], cfg["sizeKey"])),
    ("write", "openpyxl"): lambda cfg: write_openpyxl(cfg["workload"], cfg["rows"]),
}


def run_worker(cfg):
    fn = OPS.get((cfg["op"], cfg["lib"]))
    if fn is None:
        print(json.dumps({"ok": False, "reason": "no such op"}))
        return
    for _ in range(cfg["warmup"]):
        fn(cfg)
    times = []
    for _ in range(cfg["iters"]):
        t0 = time.perf_counter()
        fn(cfg)
        times.append((time.perf_counter() - t0) * 1000)
    print(json.dumps({
        "ok": True, "lib": cfg["lib"], "op": cfg["op"], "workload": cfg["workload"],
        "sizeKey": cfg["sizeKey"], "timeMs": statistics.median(times), "peakRssMB": peak_rss_mb(),
    }))


def orchestrate(size_keys, workloads):
    runners = [("read", "openpyxl"), ("read", "python-calamine"), ("write", "openpyxl")]
    results = []
    for size_key, rows in SIZES:
        if size_key not in size_keys:
            continue
        iters = 5 if size_key == "10k" else 3 if size_key == "100k" else 2
        for workload in workloads:
            for op, lib in runners:
                if op == "read" and not os.path.exists(fixture_path(workload, size_key)):
                    print(f"  · skip {op}/{lib}/{workload}/{size_key} — no fixture (run `pnpm bench` first)")
                    continue
                cfg = {"op": op, "lib": lib, "workload": workload, "sizeKey": size_key,
                       "rows": rows, "iters": iters, "warmup": 1}
                sys.stdout.write(f"▶ {op:5} · {workload:7} · {size_key:4} · {lib} … ")
                sys.stdout.flush()
                proc = subprocess.run(
                    [sys.executable, __file__, "--worker", json.dumps(cfg)],
                    capture_output=True, text=True,
                )
                line = proc.stdout.strip().splitlines()[-1] if proc.stdout.strip() else ""
                try:
                    res = json.loads(line)
                except Exception:
                    res = {"ok": False, "reason": (proc.stderr or "crash").strip().splitlines()[-1:]}
                if res.get("ok"):
                    print(f"{res['timeMs']:.0f} ms · peak {res['peakRssMB']:.0f} MB")
                    results.append(res)
                else:
                    print(f"n/a ({res.get('reason')})")
    out = {"machine": f"{sys.platform} · Python {sys.version.split()[0]}", "date": date.today().isoformat(), "results": results}
    os.makedirs(CACHE, exist_ok=True)
    with open(os.path.join(CACHE, "python.json"), "w") as f:
        json.dump(out, f, indent=2)
    print(f"\nWrote {os.path.join(CACHE, 'python.json')} ({len(results)} cells)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--worker", help="internal: run one cell from a JSON config")
    ap.add_argument("--sizes", default="10k,100k,1M")
    ap.add_argument("--workloads", default="numbers,strings,styled")
    args = ap.parse_args()
    if args.worker:
        run_worker(json.loads(args.worker))
    else:
        orchestrate(args.sizes.split(","), args.workloads.split(","))


if __name__ == "__main__":
    main()
